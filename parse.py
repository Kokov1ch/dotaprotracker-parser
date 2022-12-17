import openpyxl
from bs4 import BeautifulSoup
import requests

response = requests.get('https://dota2protracker.com/')
src = response.content
soup = BeautifulSoup(src, "lxml")
def get_heroes_data():
    roles_heroes = []
    all_roles = soup.find("div", {'class': "table-wrapper table-small top-hero-table"})
    for i in range (2, 7):
        tab_heroes_data = all_roles.find("div", {'class': f"TH tabs-{i}"}).find_all('tr')
        heroes = []

        for i in range (1, len(tab_heroes_data)):
            hero_pic = tab_heroes_data[i].find("td", {'class': "td-hero-pic"}).find('a').text.replace('\n\n', '').replace('\n', '').replace(' ', '')
            hero_winrate = tab_heroes_data[i].find("td", {'class': "td-winrate"}).find("div", {'class': "perc-wr"}).find("span").text
            hero_matches = tab_heroes_data[i].find("td", {'class': "td-matches"}).find("div", {'class': "perc-wr"}).text
            heroes.append([hero_pic, hero_winrate, hero_matches])
        roles_heroes.append(heroes)

    heroes_result = openpyxl.Workbook()
    carry = heroes_result['Sheet']
    carry.title = 'Carry'
    heroes_result.create_sheet('Mid')
    heroes_result.create_sheet('Offlane')
    heroes_result.create_sheet('Soft support')
    heroes_result.create_sheet('Hard support')

    carry = heroes_result["Carry"]
    mid = heroes_result["Mid"]
    offlane = heroes_result["Offlane"]
    soft_support = heroes_result["Soft support"]
    hard_support = heroes_result["Hard support"]

    carry.append(['Hero', 'Winrate', 'Matches'])
    mid.append(['Hero', 'Winrate', 'Matches'])
    offlane.append(['Hero', 'Winrate', 'Matches'])
    soft_support.append(['Hero', 'Winrate', 'Matches'])
    hard_support.append(['Hero', 'Winrate', 'Matches'])

    for i in roles_heroes[0]:
        carry.append([i[0], i[1], i[2]])

    for i in roles_heroes[1]:
        mid.append([i[0], i[1], i[2]])

    for i in roles_heroes[2]:
        offlane.append([i[0], i[1], i[2]])

    for i in roles_heroes[3]:
        soft_support.append([i[0], i[1], i[2]])

    for i in roles_heroes[4]:
        hard_support.append([i[0], i[1], i[2]])

    heroes_result.save("heroes.xlsx")
def get_players_data():
    roles_players = []
    all_roles = soup.find('div', {'class': 'top-players-table'})
    for i in range (2, 7):
        tab_players_data = all_roles.find("div", {'class': f"TP tabs-{i}"}).find_all('tr')
        players = []
        for i in range(1, len(tab_players_data)):
            player_pic = tab_players_data[i].find("td", {'class': "td-player"}).find('a').text
            player_matches = tab_players_data[i].find("td", {'class': "td-matches"}).find("div", {'class': "perc-wr"}).text
            player_winrate = tab_players_data[i].find("td", {'class': "td-winrate"}).find('span').text
            players.append([player_pic, player_matches, player_winrate])
        roles_players.append(players)

    players_result = openpyxl.Workbook()
    carry = players_result['Sheet']
    carry.title = 'Carry'
    players_result.create_sheet('Mid')
    players_result.create_sheet('Offlane')
    players_result.create_sheet('Soft support')
    players_result.create_sheet('Hard support')

    mid = players_result["Mid"]
    offlane = players_result["Offlane"]
    soft_support = players_result["Soft support"]
    hard_support = players_result["Hard support"]

    carry.append(['Player', 'Matches', 'Winrate'])
    mid.append(['Player', 'Matches', 'Winrate'])
    offlane.append(['Player', 'Matches', 'Winrate'])
    soft_support.append(['Player', 'Matches', 'Winrate'])
    hard_support.append(['Player', 'Matches', 'Winrate'])


    for i in roles_players[0]:
        carry.append([i[0], i[1], i[2]])

    for i in roles_players[1]:
        mid.append([i[0], i[1], i[2]])

    for i in roles_players[2]:
        offlane.append([i[0], i[1], i[2]])

    for i in roles_players[3]:
        soft_support.append([i[0], i[1], i[2]])

    for i in roles_players[4]:
        hard_support.append([i[0], i[1], i[2]])
    players_result.save("players.xlsx")
get_heroes_data()
get_players_data()
